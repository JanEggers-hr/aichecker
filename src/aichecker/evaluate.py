# evaluate.py
"""
Generisches Skript zur Auswertung der Posts mit KI. Setzt ein hydriertes
Post - also eins mit den Pfaden zu gespeicherten Medien - voraus. 

- 'text': Text des Posts
- 'media': dict oder Liste von dicts:
    - 'type': 'image' | 'video' | 'voice'
    - 'file': Pfad zum Medium

Gibt zurück: Liste posts, ergänzt:
- 'media':
    - 'description' bzw. 'transcription': KI-Bildbeschreibung bzw. Transkript
    - 'aiornot_ai_score': dict mit den KI-Check-Ergebnissen
    #   - 'score'
    #   - 'confidence'
    #   - 'generator' (list)
    - 'hive_visual': dict mit den KI-Check-Ergebnissen
        - 'ai_score': real,
        - 'most_likely_model': string,
        - 'models': [ {'class': string, 'score': real}, ... ]

- 'detectora_ai_score'
- 'aiornot_ai_max_score'
- 'hive_visual_max_score'

"""
import aiohttp
import asyncio
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import logging
import base64
import pandas as pd
import os
# Imports für den aufgehübschten XLSX-Export
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Alignment


from .transcribe import gpt4_description, transcribe, convert_mp4_to_mp3, convert_ogg_to_mp3
from .check_wrappers import detectora_wrapper, aiornot_wrapper, transcribe_async, describe_async, detectora_async, aiornot_async, hive_visual, hive_visual_async

# Routine checkt eine Post-Liste, wie sie aus den ig_post_parse Routinen kommen.
# Wenn noch kein KI-Check vorliegt, wird er ergänzt. 
# Setzt allerdings voraus, dass die entsprechenden Inhalte schon abgespeichert sind.
# Routine fragt asynchron die APIs von AIORNOT und Hive ab
# und lässt Audio-Inhalte transkribieren/Bilder beschreiben


# Die Routine kann auch nur die Texte oder nur die Bilder checken lassen.
# Wenn check_texts = False, dann wird der Text nicht gecheckt.
# Wenn check_images = False, dann werden die Bilder nicht gecheckt.

async def evaluate_async(posts: List[Dict[str, Any]], check_texts: bool = True, check_images: bool = True) -> List[Dict[str, Any]]:
    async with aiohttp.ClientSession() as session:
        tasks = []
        # Semaphore to keep it to 2 calls at a time
        semaphore = asyncio.Semaphore(2)
        semaphore2 = asyncio.Semaphore(1)
        async def hive_visual_with_delay(session, file_path, semaphore):
            async with semaphore:
                result = await hive_visual_async(session,file_path)
                await asyncio.sleep(1.1)  # Rate limit delay:
                # According to Hive, the default is one query per second (be it over
                # the sync oder async API. They have temporarily set this value to .5s for us
                # but I tend to err on the side of caution.)
                return result
            
        async def aiornot_async_with_delay(file_path, semaphore, is_image = True):
            async with semaphore:
                result = await aiornot_async(file_path, is_image = False)
                await asyncio.sleep(1.1)
                return result
         
        for post in posts:
            if check_images and 'media' in post:
                # media in Liste packen, falls Einzelelement
                if isinstance(post['media'], dict):
                    post['media'] = [post['media']]
                    was_single = True
                else:
                    was_single = False

                for m in post['media']:
                    media_type = m['type']
                    file_path = m['file']
                    # Transcription or description? Video gets transcription, image gets description
                    if media_type in ['image', 'sticker', 'photo']:
                        with open(file_path, 'rb') as f:
                            image_data = base64.b64encode(f.read()).decode('utf-8')
                        tasks.append(describe_async(f"data:image/jpeg;base64,{image_data}"))
                        tasks.append(aiornot_async(file_path, is_image=True))
                    elif media_type == 'video': # Dann ist es ein Video
                        tasks.append(transcribe_async(file_path))
                        tasks.append(aiornot_async_with_delay(convert_mp4_to_mp3(file_path), semaphore2, is_image=False))
                    elif media_type == 'voice':
                        if file_path.endswith('.oga') or file_path.endswith('.ogg'):
                            file_path = convert_ogg_to_mp3(file_path)
                        tasks.append(transcribe_async(file_path))
                        tasks.append(aiornot_async_with_delay(file_path, is_image=False))
                    # Add hive_visual_async task with rate limit
                    tasks.append(hive_visual_with_delay(session, file_path, semaphore))
                    # Da hive ein Rate-Limit von 1s hat, habe ich hier ursprünglich synchron 
                    # gelesen; inzwischen umgestellt auf Asynchron. Siehe Definition oben.                   
            if check_texts and 'text' in post:
                tasks.append(detectora_async(post.get('text','')))

        results = await asyncio.gather(*tasks)

        index = 0
        for post in posts:
            if check_images:
                aiornot_max = 0
                hive_max = 0
                for m in post['media']:
                    # Derzeit werden Medien entweder transkribiert oder beschrieben. 
                    # Denkbar ist aber auch eine Beschreibung mit Transkription. Das 
                    # wird hier vorbereitet und muss dann für den entsprechenden Medientyp
                    # oben mit angefordert werden. 
                    if m['type'] in ['image', 'sticker', 'photo']:
                        m['description'] = results[index]
                        index +=1
                    if m['type'] in ['voice', 'video']:
                        m['transcription'] = results[index]
                        index +=1
                    aiornot = results[index]
                    index +=1
                    m['aiornot_ai_score'] = aiornot
                    if aiornot is not None and aiornot_max < aiornot.get('confidence',0):
                        aiornot_max = aiornot.get('confidence',0)
                    hive = results[index]
                    m['hive_visual_ai'] = hive
                    index += 1
                    # jetzt den jeweils höchsten HiveScore nehmen
                    if hive is not None and hive_max < hive.get('ai_score',0):
                        hive_max = hive.get('ai_score',0)
                # Falls der Post nur ein einzelnes dict in media
                # enthalten hat und keine Liste: zurückverwandeln.
                if was_single:
                    post['media']=post['media'][0]
                post['aiornot_ai_max_score'] = aiornot_max
                post['hive_visual_ai_max_score'] = hive_max
            if check_texts:
                post['detectora_ai_score'] = results[index]
                index += 1
# Eingerückt für die ausgeklammerte with session Anweisung oben
        return posts

def evaluate_sync(posts, check_texts=True, check_images=True):
    # Synchrone, also LAAAAANGSAME KI-Bewertung der Medien
    # mit Detectora, aiornot, und Hive. 
    for post in posts:
        if ('detectora_ai_score' not in post) and check_texts:
            # Noch keine KI-Einschätzung für den Text?
            post['detectora_ai_score'] = detectora_wrapper(post.get('caption', ''))
        if ('aiornot_ai_score' not in post) and check_images:
            max_ai_score = 0
            if post.get('video'):
                # Alle Videos analysieren und den höchsten Score ermitteln
                for video_url in post['videos']:
                    ai_score = aiornot_wrapper(convert_mp4_to_mp3(video_url), is_image=False)
                    max_ai_score = max(max_ai_score, ai_score)
            if post.get('image'):
                # Alle Bilder analysieren und den höchsten Score ermitteln
                for image_url in post['images']:
                    ai_score = aiornot_wrapper(image_url, is_image=True)
                    max_ai_score = max(max_ai_score, ai_score)
            post['aiornot_ai_score'] = max_ai_score
    return posts

# Hilfsfunktion: Zählt die Anzahl der Medien und produziert einen Markdown-String mit der Auswertung
# Immer die gleiche Struktur: Liste mit dict, in diesem dict ein Key 'media'... dort durchzählen. 

def eval_scans(posts, t_detectora, t_aiornot, t_hive_visual):
    n_texts = 0
    n_videos = 0
    n_images = 0
    n_audios = 0
    n_ai_texts = 0
    n_ai_videos = 0
    n_ai_images = 0
    n_ai_audios = 0
    
    for post in posts:
        if post['text'] is not None:
            n_texts += 1
            # Detectora-Score für diesen Text abrufen; wenn über der Schwelle, 
            # KI-Texte um eins hochzählen
            n_ai_texts += 1 if post.get('detectora_ai_score',0) > t_detectora else 0
        for m in post['media']:
            if m.get('type') == 'video':
                n_videos += 1  
                if m.get('aiornot_ai_score') is not None:
                    aiornot = m.get('aiornot_ai_score').get('confidence')
                else:
                    aiornot = None
                hive = m.get('hive_visual',{}).get('ai_score')
                if hive is None:
                    if aiornot is not None: 
                        n_ai_videos += 1 if aiornot >= t_aiornot else 0
                elif aiornot is None:
                    n_ai_videos += 1 if hive >= t_hive_visual else 0
                else:
                    n_ai_videos +=1 if hive>= t_hive_visual and aiornot >= t_aiornot else 0
            if m.get('type') in ['sticker','image', 'photo']:
                n_images += 1  
                if m.get('aiornot_ai_score') is not None:
                    aiornot = m.get('aiornot_ai_score').get('confidence')
                else:
                    aiornot = None
                hive = m.get('hive_visual',{}).get('ai_score')
                if hive is None:
                    if aiornot is not None: 
                        n_ai_images += 1 if aiornot >= t_aiornot else 0
                elif aiornot is None:
                    n_ai_images += 1 if hive >= t_hive_visual else 0
                else:
                    n_ai_images +=1 if hive>= t_hive_visual and aiornot >= t_aiornot else 0
            if m.get('type') in ['audio', 'voice']:
                n_audios += 1
                if m.get('aiornot_ai_score') is not None:
                    aiornot = m.get('aiornot_ai_score').get('confidence')
                else:
                    aiornot = None
                hive = m.get('hive_visual',{}).get('ai_score')
                if hive is None:
                    if aiornot is not None: 
                        n_ai_images += 1 if aiornot >= t_aiornot else 0
                elif aiornot is None:
                    n_ai_images += 1 if hive >= t_hive_visual else 0
                else:
                    n_ai_images +=1 if hive>= t_hive_visual and aiornot >= t_aiornot else 0
    if n_texts > 0:
        detectora_mean = round(n_ai_texts/n_texts,2)
    else:
        detectora_mean = 0
    if n_images > 0:
        aiornot_mean = round(n_ai_images/n_images,2)
    else:
        aiornot_mean = 0
    if n_videos > 0:
        hive_mean = round(n_ai_videos/n_videos,2)
    else:
        hive_mean = 0
    return {'n_texts': n_texts, 
            'n_ai_texts': n_ai_texts, 
            'n_images': n_images, 
            'n_ai_images': n_ai_images, 
            'n_videos': n_videos, 
            'n_ai_videos': n_ai_videos, 
            'n_audios': n_audios, 
            'n_ai_audios': n_ai_audios, 
            'detectora_mean': detectora_mean, 
            'aiornot_mean': aiornot_mean, 
            'hive_mean': hive_mean}

def evaluate_scans(posts, t_detectora, t_aiornot, t_hive_visual):
    n_posts = len(posts)
    e_dict = eval_scans(posts, t_detectora, t_aiornot, t_hive_visual)
    
    eval_str = f'**Analysierte Posts:** {n_posts}\n\n'
    eval_str += f"- **Texte über der KI-Schwelle von {t_detectora}:** {e_dict['n_ai_texts']} von {e_dict['n_texts']}\n"
    eval_str += f"- **Bilder über der KI-Schwelle von {t_aiornot}:** {e_dict['n_ai_images']} von {e_dict['n_images']}\n"
    eval_str += f"- **Videos mit KI-verdächtigem Audio/Video:** {e_dict['n_ai_videos']} von {e_dict['n_videos']}\n"
    eval_str += f"- **Voice Messages mit KI-verdächtigem Audio:** {e_dict['n_ai_audios']} von {e_dict['n_audios']}\n"
    return eval_str

def most_likely_aiornot_model(aiornot):
    # Liest die Generator-Keys aus und gibt das wahrscheinlichste zurück
    generator = aiornot.get('generator', {})
    model_str=""
    for model in list(generator.keys()):
        if generator[model].get('is_detected'):
            model_str += f"{model} {generator[model].get('confidence')*100:.1f}% "
    return model_str
        

def export_to_xlsx(posts, filename):
    # Schreibt alle Spalten in eine Excel-Datei, die keine Objekte enthalten. Ausnahme: 
    # Media wird "explodiert", d.h. neue Zeilen für jedes Medien-Objekt mit:
    # type, url, description, aiornot_score, aiornot_guess, hive_score, hive_guess
    fname = os.path.splitext(filename)[0]+".xlsx"
    export_posts = []
    for post in posts:
        media = post.get('media', [])
        post.pop('media')
        post.pop('aiornot_ai_max_score')
        post.pop('hive_visual_ai_max_score')
        # Drop all keys containing objects
        for key in list(post.keys()):
            if isinstance(post[key], dict):
                post.pop(key)

        # Explode media
        for m in media:
            post['media_type'] = m['type']
            post['file'] = m['file']
            if m.get('description'):
                post['description'] = m['description']
            if m.get('transcription'):
                post['transcription'] = m['transcription']
            aiornot = m.get('aiornot_ai_score')
            if aiornot:
                post['aiornot_score'] = aiornot.get('confidence',None)
                post['aiornot_guess'] =most_likely_aiornot_model(aiornot)
            hive = m.get('hive_visual_ai')
            if hive:
                post['hive_score'] = hive.get('ai_score')
                post['hive_guess'] = hive.get('most_likely_model')
            export_posts.append(post.copy())
        if 'media' in post:
            media = post.get('media')
            for m in media:
                post
        else:
            export_posts.append(post)

    df = pd.DataFrame(posts)
        # Export als Workbook mit eingestellten Spaltenbreiten
    try:
        # Create a Workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Append the DataFrame to the worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Set column width
        text_columns = {
            'A': 20, # id
            'B': 40, # text
            'C': 10, # type
            'D': 20, # timestamp
            'E': 6, # detectora
            'F': 6, # media_type
            'G': 12, # file URL
            'H': 6, # aiornot score
            'I': 12, # aiornot guess
            'J': 6, # hive score
            'K': 12, # hive guess
        }
        
        for col, width in text_columns.items():
            ws.column_dimensions[col].width = width

        # Apply text wrapping to the second column
        def apply_text_wrapping(ws, column_index):
            for row in ws.iter_rows(min_col=column_index, max_col=column_index, min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

        # Apply text wrapping to the second column (column index 2)
        # apply_text_wrapping(ws, 2)

        # Save the workbook
        wb.save(fname)
        logging.debug(f"XLSX-Tabelle gespeichert: {fname}")
        return(fname)

    except Exception as e:
        logging.error(f"Fehler beim Anlegen von {fname}: {e}")
        df.to_excel(fname, index=False)
    return fname